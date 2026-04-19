"""
exporter/excel_exporter.py

Exports processed B2B leads to a richly-formatted Excel workbook.

Sheets produced:
  1. All Leads          — every unique validated lead
  2. No Website Leads   — leads missing a website URL
  3. Leads With Phone   — leads that have a normalised phone number

Formatting:
  • Frozen header row with bold, white-on-dark-blue text
  • Alternating row shading (white / light blue)
  • Auto-fitted column widths
  • Excel auto-filter on every sheet
  • Hyperlinks in the 'website' column where available
  • Summary stats in a dedicated 'Summary' sheet
"""

import os
import logging
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

log = logging.getLogger(__name__)

# ─── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # Dark navy
ROW_ALT_FILL  = PatternFill("solid", start_color="DCE6F1")   # Light blue
ROW_EVEN_FILL = PatternFill("solid", start_color="FFFFFF")   # White
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
LINK_FONT     = Font(name="Arial", size=10, color="0563C1", underline="single")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

THIN_BORDER_SIDE = Side(style="thin", color="B8CCE4")
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)

COLUMNS = ["name", "company", "phone", "email", "website", "city", "source"]
HEADERS = ["Name", "Company", "Phone", "Email", "Website / LinkedIn", "City", "Source"]

# Minimum and maximum column widths (characters)
COL_MIN_WIDTH = 14
COL_MAX_WIDTH = 55


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _set_header_row(ws: Worksheet) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = CELL_BORDER

    ws.row_dimensions[1].height = 22


def _write_data_rows(ws: Worksheet, leads: list[dict[str, Any]]) -> None:
    """Write lead data rows with alternating fill and hyperlinks."""
    for row_idx, lead in enumerate(leads, start=2):
        fill = ROW_EVEN_FILL if row_idx % 2 == 0 else ROW_ALT_FILL

        for col_idx, field in enumerate(COLUMNS, start=1):
            value = lead.get(field, "") or ""
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = LEFT_ALIGN
            cell.border    = CELL_BORDER

            # Hyperlink for website / LinkedIn URL
            if field == "website" and value.startswith("http"):
                cell.hyperlink = value
                cell.font = LINK_FONT
            else:
                cell.font = DATA_FONT


def _auto_fit_columns(ws: Worksheet) -> None:
    """Set column widths based on the longest cell content in each column."""
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = len(HEADERS[col_idx - 1])
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, COL_MIN_WIDTH), COL_MAX_WIDTH
        )


def _finalise_sheet(ws: Worksheet, leads: list[dict[str, Any]]) -> None:
    """Apply all styling and data to a worksheet."""
    _set_header_row(ws)
    _write_data_rows(ws, leads)
    _auto_fit_columns(ws)
    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions
    # Freeze the header row
    ws.freeze_panes = "A2"


def _add_summary_sheet(wb: openpyxl.Workbook, all_leads: list[dict]) -> None:
    """Add a 'Summary' sheet with run statistics."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "2E75B6"

    gm_leads = [l for l in all_leads if l.get("source") == "google_maps"]
    li_leads = [l for l in all_leads if l.get("source") == "linkedin"]
    no_web   = [l for l in all_leads if not l.get("website")]
    with_ph  = [l for l in all_leads if l.get("phone")]

    rows = [
        ("B2B Lead Generation — Run Summary", ""),
        ("", ""),
        ("Generated at",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Metric",              "Count"),
        ("Total Unique Leads",  len(all_leads)),
        ("Google Maps Leads",   len(gm_leads)),
        ("LinkedIn Leads",      len(li_leads)),
        ("Leads Without Website", len(no_web)),
        ("Leads With Phone",    len(with_ph)),
    ]

    TITLE_FONT   = Font(name="Arial", bold=True, size=14, color="1F3864")
    LABEL_FONT   = Font(name="Arial", bold=True, size=10)
    VALUE_FONT   = Font(name="Arial", size=10)
    HEADER2_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    HDR2_FILL    = PatternFill("solid", start_color="2E75B6")

    for r_idx, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_b = ws.cell(row=r_idx, column=2, value=value)

        if r_idx == 1:
            cell_a.font = TITLE_FONT
        elif label == "Metric":
            cell_a.font  = HEADER2_FONT
            cell_b.font  = HEADER2_FONT
            cell_a.fill  = HDR2_FILL
            cell_b.fill  = HDR2_FILL
            cell_a.alignment = CENTER_ALIGN
            cell_b.alignment = CENTER_ALIGN
        elif label:
            cell_a.font = LABEL_FONT
            cell_b.font = VALUE_FONT
            cell_b.alignment = CENTER_ALIGN

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A1"


# ─── Public API ───────────────────────────────────────────────────────────────

def export_to_excel(leads: list[dict[str, Any]], output_path: str | None = None) -> str:
    """
    Write leads to an Excel workbook with three data sheets + summary.

    Args:
        leads:       Deduplicated, validated lead dicts.
        output_path: Full path to write the .xlsx file.
                     Defaults to config.OUTPUT_DIR / config.OUTPUT_FILENAME.

    Returns:
        Absolute path of the saved file.
    """
    if output_path is None:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(config.OUTPUT_DIR, config.OUTPUT_FILENAME)

    # ── Segment leads ─────────────────────────────────────────────────────
    gm_leads  = [l for l in leads if l.get("source") == "google_maps"]
    li_leads  = [l for l in leads if l.get("source") == "linkedin"]
    no_web    = [l for l in leads if not (l.get("website") or "").strip()]
    with_ph   = [l for l in leads if (l.get("phone") or "").strip()]

    log.info(
        "[Exporter] Segments — All: %d  |  Google Maps: %d  |  "
        "LinkedIn: %d  |  No Website: %d  |  With Phone: %d",
        len(leads), len(gm_leads), len(li_leads), len(no_web), len(with_ph),
    )

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: All Leads
    ws_all = wb.active
    ws_all.title = "All Leads"
    ws_all.sheet_properties.tabColor = "1F3864"
    _finalise_sheet(ws_all, leads)

    # Sheet 2: No Website Leads
    ws_no_web = wb.create_sheet("No Website Leads")
    ws_no_web.sheet_properties.tabColor = "C55A11"
    _finalise_sheet(ws_no_web, no_web)

    # Sheet 3: Leads With Phone
    ws_phone = wb.create_sheet("Leads With Phone")
    ws_phone.sheet_properties.tabColor = "375623"
    _finalise_sheet(ws_phone, with_ph)

    # Sheet 4: Summary
    _add_summary_sheet(wb, leads)

    # ── Save ──────────────────────────────────────────────────────────────
    try:
        wb.save(output_path)
        log.info("[Exporter] ✓ Workbook saved → %s", os.path.abspath(output_path))
    except Exception as exc:
        log.error("[Exporter] Failed to save workbook: %s", exc)
        raise

    return os.path.abspath(output_path)
